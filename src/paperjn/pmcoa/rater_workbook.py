from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd

from paperjn.pmcoa.rater_validation import SYMPTOM_TAGS_V1
from paperjn.utils.paths import ensure_dir


@dataclass(frozen=True)
class RaterWorkbookOutputs:
    out_dir: Path
    rater1_xlsx: Path
    rater2_xlsx: Path


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _build_workbook(
    *,
    out_xlsx: Path,
    template_df: pd.DataFrame,
    packet_texts: list[str],
    instructions_md: str,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("Missing dependency: openpyxl") from exc

    wb = Workbook()
    ws_instr = wb.active
    ws_instr.title = "INSTRUCTIONS"
    ws = wb.create_sheet("RATINGS")

    # Instructions (one line per row for easy reading in Google Sheets).
    ws_instr["A1"] = "Clinician Rating Instructions (Blinded)"
    ws_instr["A1"].font = Font(bold=True, size=14)
    ws_instr["A3"] = "Copy/paste of rubric:"
    ws_instr["A3"].font = Font(bold=True)
    start_row = 5
    for i, line in enumerate(instructions_md.splitlines(), start=start_row):
        ws_instr.cell(row=i, column=1, value=line)
    ws_instr.column_dimensions["A"].width = 120
    ws_instr.freeze_panes = "A5"

    # Ratings sheet
    df = template_df.copy()
    df.insert(4, "packet_text", packet_texts)

    # Write header
    header_fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col))
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            v = "" if pd.isna(value) else value
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Column widths (Google Sheets will approximate)
    widths: dict[str, float] = {
        "case_uid": 10,
        "pmcid": 12,
        "case_id": 14,
        "packet_path": 18,
        "packet_text": 120,
        "text_adequate": 12,
        "is_ftld_spectrum": 14,
        "is_ppa": 8,
        "notes": 40,
    }
    for tag in SYMPTOM_TAGS_V1:
        widths[f"tag__{tag}"] = 16
    for col_idx, col in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = float(widths.get(str(col), 14))

    # Data validation (0/1 dropdown) for rating columns.
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True, showDropDown=True)
    ws.add_data_validation(dv)
    rating_cols = ["text_adequate", "is_ftld_spectrum", "is_ppa"] + [f"tag__{t}" for t in SYMPTOM_TAGS_V1]
    for col_name in rating_cols:
        if col_name not in df.columns:
            continue
        cidx = int(df.columns.get_loc(col_name)) + 1
        col_letter = get_column_letter(cidx)
        dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def make_rater_workbooks(
    *,
    rater_sample_dir: Path,
    out_dir: Path | None = None,
) -> RaterWorkbookOutputs:
    """Create Google-Sheets-friendly Excel workbooks (one per rater) with embedded packet text."""
    rater_sample_dir = Path(rater_sample_dir).resolve()
    share_dir = rater_sample_dir / "share"
    if not share_dir.exists():
        raise FileNotFoundError(f"Missing share/ directory: {share_dir}")

    r1_csv = share_dir / "rater1_template.csv"
    r2_csv = share_dir / "rater2_template.csv"
    instr_md_path = share_dir / "RATER_INSTRUCTIONS.md"
    if not r1_csv.exists() or not r2_csv.exists():
        raise FileNotFoundError("Missing rater templates in share/: rater1_template.csv and/or rater2_template.csv")
    if not instr_md_path.exists():
        raise FileNotFoundError(f"Missing RATER_INSTRUCTIONS.md: {instr_md_path}")

    if out_dir is None:
        out_dir = share_dir
    out_dir = ensure_dir(Path(out_dir).resolve())

    instructions_md = _read_text(instr_md_path)

    def _load_with_packet_text(csv_path: Path) -> tuple[pd.DataFrame, list[str]]:
        df = pd.read_csv(csv_path)
        if "packet_path" not in df.columns:
            raise ValueError(f"Missing packet_path column: {csv_path}")
        texts: list[str] = []
        for rel in df["packet_path"].astype(str).tolist():
            p = (share_dir / rel).resolve()
            if not p.exists():
                raise FileNotFoundError(f"Missing packet file referenced by template: {rel}")
            texts.append(_read_text(p))
        return df, texts

    df1, texts1 = _load_with_packet_text(r1_csv)
    df2, texts2 = _load_with_packet_text(r2_csv)

    r1_xlsx = out_dir / "HUMAN_RATER1_WORKBOOK.xlsx"
    r2_xlsx = out_dir / "HUMAN_RATER2_WORKBOOK.xlsx"

    _build_workbook(out_xlsx=r1_xlsx, template_df=df1, packet_texts=texts1, instructions_md=instructions_md)
    _build_workbook(out_xlsx=r2_xlsx, template_df=df2, packet_texts=texts2, instructions_md=instructions_md)

    return RaterWorkbookOutputs(out_dir=out_dir, rater1_xlsx=r1_xlsx, rater2_xlsx=r2_xlsx)

